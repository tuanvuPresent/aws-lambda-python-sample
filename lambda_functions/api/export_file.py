import base64
import io

import openpyxl


def lambda_handler(event, context):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws['A1'] = 'Name'
    ws['B1'] = 'Age'
    ws['A2'] = 'John'
    ws['B2'] = 30
    ws['A3'] = 'Jane'
    ws['B3'] = 25

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    file_base64 = base64.b64encode(file_stream.getvalue()).decode('utf-8')

    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename="output.xlsx"'
        },
        'body': file_base64,
        'isBase64Encoded': True
    }
